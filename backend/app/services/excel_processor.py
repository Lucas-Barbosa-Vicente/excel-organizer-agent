import re
import asyncio
import logging
from typing import List, Optional
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

from app.schemas.organize import OrganizeParameters, SortRule, ColorRule
from app.utils.colors import COLOR_MAP

logger = logging.getLogger(__name__)

MAX_COLOR_RULES = 10
MAX_CATEGORIES = 50
LARGE_FILE_BYTES = 10 * 1024 * 1024  # 10 MB


class ExcelProcessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.dataframes: dict[str, pd.DataFrame] = {}
        self.transformations: list[str] = []

    async def load(self):
        file_size = await asyncio.to_thread(self._get_file_size)
        self.workbook = await asyncio.to_thread(
            openpyxl.load_workbook, self.file_path
        )
        if file_size > LARGE_FILE_BYTES:
            await asyncio.to_thread(self._load_large_file)
        else:
            await asyncio.to_thread(self._load_normal)

    def _get_file_size(self) -> int:
        import os
        return os.path.getsize(self.file_path)

    def _load_normal(self):
        with pd.ExcelFile(self.file_path) as xl:
            for sheet in xl.sheet_names:
                self.dataframes[sheet] = xl.parse(sheet)

    def _load_large_file(self):
        with pd.ExcelFile(self.file_path) as xl:
            with ThreadPoolExecutor() as ex:
                futures = {
                    sheet: ex.submit(xl.parse, sheet)
                    for sheet in xl.sheet_names
                }
                for sheet, fut in futures.items():
                    self.dataframes[sheet] = fut.result()

    def detect_customizations(self) -> list[str]:
        """Scan workbook for pre-existing colors and formulas. Returns list of findings."""
        findings = []
        try:
            for ws in self.workbook.worksheets:
                has_color = False
                has_formula = False
                for row in ws.iter_rows():
                    for cell in row:
                        if not has_color:
                            fill = cell.fill
                            if fill and fill.fill_type and fill.fill_type != "none":
                                has_color = True
                        if not has_formula:
                            if isinstance(cell.value, str) and cell.value.startswith("="):
                                has_formula = True
                        if has_color and has_formula:
                            break
                    if has_color and has_formula:
                        break
                if has_color:
                    findings.append(f"Cores de preenchimento na aba '{ws.title}'")
                if has_formula:
                    findings.append(f"Fórmulas existentes na aba '{ws.title}'")
        except Exception as e:
            logger.warning("Erro ao detectar customizações: %s", e)
        return findings

    async def apply_parameters(self, params: OrganizeParameters) -> dict:
        sheet_name = self.workbook.active.title
        df = self.dataframes.get(sheet_name, pd.DataFrame())
        rows_before = len(df)
        sheets_created: list[str] = []

        if params.remove_duplicates:
            df, removed = self._remove_duplicates(
                df, params.duplicate_columns, params.keep_duplicate or "first"
            )
            self.transformations.append(
                f"Duplicatas removidas: {removed} linha(s)"
            )

        if params.standardize_text:
            df = self._standardize_text(df, params.standardize_text)
            self.transformations.append(
                f"Texto padronizado em: {list(params.standardize_text.keys())}"
            )

        if params.sort_by:
            df = self._sort(df, params.sort_by)
            cols = [f"{r.column} ({r.direction})" for r in params.sort_by]
            self.transformations.append(f"Ordenado por: {', '.join(cols)}")

        self.dataframes[sheet_name] = df

        if params.color_rules:
            rules = params.color_rules[:MAX_COLOR_RULES]
            ws = self.workbook.active
            self._apply_color_rules(ws, df, rules)
            self.transformations.append(
                f"Regras de cor aplicadas: {len(rules)}"
            )

        if params.split_by_category:
            sheets_created = self._split_by_category(
                df,
                params.split_by_category,
                params.keep_original_sheet if params.keep_original_sheet is not None else True,
                params.create_summary_sheet or False,
            )
            self.transformations.append(
                f"Abas criadas por '{params.split_by_category}': {len(sheets_created)}"
            )

        return {
            "rows_before": rows_before,
            "rows_after": len(df),
            "sheets_created": sheets_created,
        }

    def _sort(self, df: pd.DataFrame, rules: List[SortRule]) -> pd.DataFrame:
        columns = [r.column for r in rules if r.column in df.columns]
        ascending = [r.direction == "asc" for r in rules if r.column in df.columns]
        if not columns:
            return df
        return df.sort_values(by=columns, ascending=ascending, na_position="last")

    def _remove_duplicates(
        self,
        df: pd.DataFrame,
        columns: Optional[List[str]],
        keep: str,
    ) -> tuple[pd.DataFrame, int]:
        before = len(df)
        subset = [c for c in (columns or []) if c in df.columns] or None
        df = df.drop_duplicates(subset=subset, keep=keep)
        return df, before - len(df)

    def _standardize_text(
        self, df: pd.DataFrame, rules: dict[str, str]
    ) -> pd.DataFrame:
        ops = {
            "upper": lambda s: s.str.upper(),
            "lower": lambda s: s.str.lower(),
            "capitalize": lambda s: s.str.title(),
            "strip": lambda s: s.str.strip(),
        }
        for col, op in rules.items():
            if col not in df.columns:
                logger.warning("Coluna '%s' não encontrada para padronização.", col)
                continue
            if op not in ops:
                continue
            try:
                df[col] = ops[op](df[col].astype(str))
            except Exception as e:
                logger.warning("Erro ao padronizar coluna '%s': %s", col, e)
        return df

    def _apply_color_rules(
        self, sheet, df: pd.DataFrame, rules: List[ColorRule]
    ):
        header_offset = 2  # row 1 = header, data starts at row 2

        for rule in rules:
            if rule.column not in df.columns:
                logger.warning("Coluna '%s' não encontrada para colorir.", rule.column)
                continue
            hex_color = COLOR_MAP.get(rule.color, "FFFF00")
            fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
            col_idx = df.columns.get_loc(rule.column) + 1  # 1-based

            for row_idx, value in enumerate(df[rule.column], start=header_offset):
                try:
                    if self._match_rule(value, rule):
                        for cell in sheet[row_idx]:
                            cell.fill = fill
                except Exception as e:
                    logger.warning("Erro ao aplicar regra de cor na linha %d: %s", row_idx, e)

    def _match_rule(self, value, rule: ColorRule) -> bool:
        op = rule.operator
        rv = rule.value

        str_val = "" if pd.isna(value) else str(value)

        if op == "is_empty":
            return str_val.strip() == ""
        if op == "is_not_empty":
            return str_val.strip() != ""
        if rv is None:
            return False
        if op == "equals":
            return str_val == rv
        if op == "not_equals":
            return str_val != rv
        if op == "starts_with":
            return str_val.startswith(rv)
        if op == "ends_with":
            return str_val.endswith(rv)
        if op == "contains":
            return rv in str_val
        if op == "greater_than":
            try:
                return float(str_val) > float(rv)
            except ValueError:
                return str_val > rv
        if op == "less_than":
            try:
                return float(str_val) < float(rv)
            except ValueError:
                return str_val < rv
        return False

    def _split_by_category(
        self,
        df: pd.DataFrame,
        column: str,
        keep_original: bool,
        create_summary: bool,
    ) -> list[str]:
        if column not in df.columns:
            logger.warning("Coluna '%s' não encontrada para dividir.", column)
            return []

        categories = df[column].dropna().unique()[:MAX_CATEGORIES]
        existing_names: set[str] = set(self.workbook.sheetnames)
        created: list[str] = []

        for cat in categories:
            safe_name = self._safe_sheet_name(str(cat), existing_names)
            existing_names.add(safe_name)
            created.append(safe_name)

            subset = df[df[column] == cat].reset_index(drop=True)
            ws = self.workbook.create_sheet(title=safe_name)
            self._write_df_to_sheet(ws, subset)

        if create_summary:
            summary_name = self._safe_sheet_name("Resumo", existing_names)
            ws_summary = self.workbook.create_sheet(title=summary_name)
            summary_df = (
                df[column].value_counts().reset_index()
            )
            summary_df.columns = [column, "Contagem"]
            self._write_df_to_sheet(ws_summary, summary_df)
            created.append(summary_name)

        if not keep_original:
            active_title = self.workbook.active.title
            if active_title in self.workbook.sheetnames:
                del self.workbook[active_title]

        return created

    def _safe_sheet_name(self, name: str, existing: set[str]) -> str:
        name = re.sub(r'[\\/?*\[\]]', "_", name)[:31]
        if name not in existing:
            return name
        for i in range(1, 1000):
            candidate = f"{name[:28]}_{i}"
            if candidate not in existing:
                return candidate
        return name

    def _write_df_to_sheet(self, ws, df: pd.DataFrame):
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))

    async def save(self, output_path: str):
        active_title = self.workbook.active.title
        if active_title in self.dataframes:
            ws = self.workbook.active
            df = self.dataframes[active_title]
            ws.delete_rows(2, ws.max_row)
            for row in df.itertuples(index=False):
                ws.append(list(row))
        await asyncio.to_thread(self.workbook.save, output_path)
        self.workbook.close()
